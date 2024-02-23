# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import pandas as pd
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'Wage Register'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["JM Frictech India Pvt Ltd","","","","","","","","","","","","","","","","","","","","","",""])
    ws.append(["Plant Name: "+ args.branch + " / " + "Contractor: " + "" + args.contractor,"","","","","","","","","","","","","","","","","","","","",""])
    ws.append(["Wage Register","","","","","","","","","","","","","","","","","","","","","",""])
    ws.append(["Payroll Date: "+ args.from_date + "" + "  to  " + "" + args.to_date,"","","","","","","","","","","","","","","","","","","","",""])
    
    if args.contractor == "UPDATER SERVICES (P) LTD":
        ws.append(["Si No","Employee ID","NAME","Payment Days","Sum of Overtime to mandays","Per day Rate","Total Basic","Total Dearness Allowance","Gross1","Bonus","Total Travel Allowance","Gross2","Employer Employees State Insurance",'Employer Provident Fund',"Service Charge","Shoes and Canteen","Billed Amount"])
    else:
        ws.append(["Si No","Employee ID","NAME","Payment Days","Sum of Overtime","Per day Basic","Per day Dearness Allowance","Per day Travel Allowance Amount","Total Basic","Total Dearness Allowance","Gross1","Bonus","Total Travel Allowance","Gross2","Employer Employees State Insurance",'Employer Provident Fund',"Service Charge","Shoes and Canteen","Billed Amount"])

    if args.branch:
        salary_slips = frappe.get_all("Salary Slip",{'start_date':args.from_date,'end_date':args.to_date,"branch":args.branch,"contractor":args.contractor},['*']) 
    i=1    
    # frappe.log_error(salary_slips)
    for ss in salary_slips:    
        if args.contractor == "UPDATER SERVICES (P) LTD":
            # frappe.log_error(ss)
            basic = round(frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':ss.name },['amount']) or 0)
            per_day_amount = 707
            base_per_day = 255
            da_per_day = 237
            ot_per_day = 54
            emp = frappe.get_doc('Employee',{"employee":ss.employee},['*'])
            da = round(frappe.get_value('Salary Detail',{'salary_component':'Dearness Allowances','parent':ss.name},['amount']) or 0) or 0
            eesic = round(frappe.get_value('Salary Detail',{'salary_component':'Employer Employees State Insurance','parent':ss.name},['amount']) or 0) or 0
            epf = round(frappe.get_value('Salary Detail',{'salary_component':'Employer Provident Fund','parent':ss.name},['amount']) or 0) or 0
            bonus = round(frappe.get_value('Salary Detail',{'salary_component':'Bonus','parent':ss.name},['amount']) or 0) or 0
            service_charge = round(frappe.get_value('Salary Detail',{'salary_component':'Service Charge','parent':ss.name},['amount']) or 0) or 0
            ot = round(frappe.get_value('Salary Detail',{'salary_component':'Over Time','parent':ss.name},['amount']) or 0) or 0
            uniform_and_shoes = round(frappe.get_value('Salary Detail',{'salary_component':'uniform and shoes','parent':ss.name},['amount']) or 0) or 0
            # ctc_c = frappe.get_value('Salary Detail',{'salary_component':'CTC Component','parent':ss.name},['amount']) or 0
            # ctc = frappe.get_value('Salary Detail',{'salary_component':'CTC','parent':ss.name},['amount']) or 0
            hra = round(frappe.get_value('Salary Detail',{'salary_component':'HRA1','parent':ss.name},['amount']) or 0)or 0
            
            gross = round(basic + da + eesic + epf + bonus + service_charge + ot + hra + uniform_and_shoes)
            gross1 = (basic + da)
            gross2 = (ot + bonus)
            ws.append([i,emp.employee,emp.employee_name,ss.payment_days,ss.overtime_hours,per_day_amount,basic,da,gross1,bonus,ot,gross2,eesic,epf,service_charge,uniform_and_shoes,gross])
        
        else:
            locale.setlocale(locale.LC_ALL, 'en_US.utf8')
            # frappe.log_error(ss)
            basic = frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':ss.name },['amount']) or 0
            per_day_amount = 707
            base_per_day = 255
            da_per_day = 237
            ot_per_day = 54
            emp = frappe.get_doc('Employee',{"employee":ss.employee},['*'])
            da = frappe.get_value('Salary Detail',{'salary_component':'Dearness Allowances','parent':ss.name},['amount']) or 0
            eesic = frappe.get_value('Salary Detail',{'salary_component':'Employer Employees State Insurance','parent':ss.name},['amount']) or 0
            epf = frappe.get_value('Salary Detail',{'salary_component':'Employer Provident Fund','parent':ss.name},['amount']) or 0
            bonus = frappe.get_value('Salary Detail',{'salary_component':'Bonus','parent':ss.name},['amount']) or 0
            service_charge = frappe.get_value('Salary Detail',{'salary_component':'Service Charge','parent':ss.name},['amount']) or 0
            ot = frappe.get_value('Salary Detail',{'salary_component':'Over Time','parent':ss.name},['amount']) or 0
            uniform_and_shoes = frappe.get_value('Salary Detail',{'salary_component':'uniform and shoes','parent':ss.name},['amount']) or 0
            # ctc_c = frappe.get_value('Salary Detail',{'salary_component':'CTC Component','parent':ss.name},['amount']) or 0
            # ctc = frappe.get_value('Salary Detail',{'salary_component':'CTC','parent':ss.name},['amount']) or 0
            hra = frappe.get_value('Salary Detail',{'salary_component':'HRA1','parent':ss.name},['amount']) or 0
    
            gross = (basic + da + eesic + epf + bonus + service_charge + ot) - uniform_and_shoes
            gross1 = (basic + da)
            gross2 = (ot + bonus)
            formatted_value = locale.format_string("%0.2f", gross, grouping=True)
            frappe.log_error(formatted_value)
            ws.append([i,emp.employee,emp.employee_name,ss.payment_days,ss.overtime_hours,base_per_day,da_per_day,ot_per_day,basic,da,gross1,bonus,ot,gross2,eesic,epf,service_charge,uniform_and_shoes,formatted_value])
        i=1+i
   
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=19)
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=19)
        ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=19)
        ws.merge_cells(start_row=4,start_column=1,end_row=4,end_column=19)

        # ws.merge_cells(start_row=8,start_column=7,end_row=8,end_column=15)
        # ws.merge_cells(start_row=8,start_column=16,end_row=8,end_column=23)

        bold_font = Font(bold=True)
        for cell in ws["1:1"]:
            cell.font = bold_font
        for cell in ws["2:2"]:
            cell.font = bold_font
        for cell in ws["3:3"]:
            cell.font = bold_font
        for cell in ws["4:4"]:
            cell.font = bold_font
        for cell in ws["5:5"]:
            cell.font = bold_font

        for rows in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=19):
            for cell in rows:
                cell.fill = PatternFill(fgColor="66a3ff", fill_type = "solid")

        for rows in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=19):
            for cell in rows:
                cell.fill = PatternFill(fgColor="ff3333", fill_type = "solid")

        for rows in ws.iter_rows(min_row=6,max_col=19):
            for cell in rows:
                cell.fill = PatternFill(fgColor="ffebe6", fill_type = "solid")
        
        border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'))
    for rows in ws.iter_rows(min_row=1, min_col=1, max_col=19):

        for cell in rows:
            cell.border = border
        

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'